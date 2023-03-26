import openpyxl as exl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    #loading the workbook
    wb = exl.load_workbook(filename)
    sheet = wb['Sheet1']

    #iterating over the rows, fixing the prices, 
    for row in range(2,sheet.max_row + 1):
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    #Select the values to add a chart
    values = Reference(sheet, min_row=2, max_row= sheet.max_row, min_col = 4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    #save the workbook
    wb.save(filename)

